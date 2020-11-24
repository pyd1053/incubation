import openpyxl

# 打开excel
wb = openpyxl.load_workbook('C:/Users/Petter/Desktop/彭寅冬/疫情期间工作/20200202/承运公司/余额清单_20200102010000.xlsx')
# 获取第一个sheet页
ws = wb[wb.sheetnames[0]]
# 显示sheet名，表行数，表列数
print("work sheet title:", ws.title)
print("work sheet rows:", ws.max_row)
print("work sheet cols:", ws.max_column)

# 建立存储数据的字典
data = {}

# 获取表格所有值
# 法1：
for i in range(0, ws.max_row):
    every_row_values = []
    every_row_cell_list = list(ws.rows)[i]
    for cell in every_row_cell_list:
        every_row_values.append(cell.value)
    data[i+1] = every_row_values
print(data)
# 法2：
for row in ws.rows:
    line = [cell.value for cell in row]
    line.insert(0, '2020-01-02')
    print(line)

# 获取某个区间的值
# 法1：
data1 = {}
for i in range(2, ws.max_row - 1):
    every_row_values = []
    every_row_cell_list = list(ws.rows)[i]
    for cell in every_row_cell_list:
        every_row_values.append(cell.value)
    data1[i+1] = every_row_values
print(data1)
# 法2：
for row_cell in ws['A3':'C26']:
    every_row_value = []
    for cell in row_cell:
        every_row_value.append(cell.value)
    every_row_value.insert(0, '2020-01-02')
    print(every_row_value)
