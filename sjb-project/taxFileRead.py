import openpyxl
import os
import pymysql

def insertIntoDB(filepath):
    print("now is read file name is:", filepath)
    a, b = os.path.split(filepath)
    # 打开excel
    wb = openpyxl.load_workbook(filepath)
    # 获取第一个sheet页
    ws = wb[wb.sheetnames[0]]

    conn = pymysql.connect('192.168.5.61', 'root', '123456', 'pyd')
    cursor = conn.cursor()

    date = b[5:9]+'-'+b[9:11]+'-'+b[11:13]

    for row_cell in ws['A3':'C27']:
        every_row_value = []
        for cell in row_cell:
            every_row_value.append(cell.value.replace(',', ''))
        sql = """INSERT INTO tax_company_bal_info(create_date, agent_id, company_name, bal_amount) VALUES(%s, %s, %s, %s)"""
        every_row_value.insert(0, date)
        cursor.execute(sql, every_row_value)

    # 涉及写操作要注意提交
    conn.commit()
    # 关闭连接
    cursor.close()
    conn.close()

#insertIntoDB('C:\\Users\\Petter\\Desktop\\彭寅冬\\疫情期间工作\\20200202\\承运公司\\余额清单_20200330010000.xlsx')